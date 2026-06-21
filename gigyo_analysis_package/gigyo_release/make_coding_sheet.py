#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_coding_sheet.py — '수 있다' 수작업 코딩용 엑셀 워크시트 생성

255개 '~ㄹ 수 있다' 용례를 사람이 직접 완화/비완화 판정하도록 만든다.
판정을 드롭다운으로 입력하면 [집계] 시트에서 BHR이 엑셀 수식으로 실시간 재계산된다.
상수(강조·비수있다완화·어절 수)는 gigyo 파이프라인에서 그대로 가져오므로,
모든 용례를 '완화'로 코딩하면 R0(현행 게재값)과 정확히 일치한다.

  python make_coding_sheet.py            # → output/suitda_coding.xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from gigyo import build_corpus, config as C
from gigyo.core import BOOSTER_ENGINE, MarkerEngine
from run_bhr_sense import PAT, classify, CRITICS

KO = "맑은 고딕"
HEDGE_FULL = MarkerEngine(C.HEDGES_INV)
HEDGE_BASE = MarkerEngine({cat: [e for e in es if e["label"] != "ㄹ 수 있다"]
                           for cat, es in C.HEDGES_INV.items()})
WIN = 55
AUTOLABEL = {"epistemic": "인식(완화?)", "judgment": "판단(완화?)",
             "ability": "능력(비완화?)", "present": "제시(비완화?)", "ambiguous": "양가(?)"}


def extract(raw):
    flat = re.sub(r"\s+", " ", raw)
    rows = []
    for m in PAT.finditer(flat):
        prev = m.group(1)
        pre = flat[max(0, m.start() - 20):m.start()].strip().split(" ")
        prev2 = pre[-1] if pre else ""
        left = flat[max(0, m.start() - WIN):m.start()]
        right = flat[m.end():m.end() + WIN]
        rows.append((left, m.group(0), right, prev, classify(prev, prev2)))
    return rows


def per_critic_constants(corpus):
    out = {}
    for c in CRITICS:
        raw = corpus.critic_raw_text(c)
        eoj = corpus.critic_eojeols(c)
        morphs = corpus.tokenizer.morphs(raw)
        booster = BOOSTER_ENGINE.count_all(morphs, raw)["강조"]["count"]
        h_full = HEDGE_FULL.count_all(morphs, raw)["완화"]["count"]
        h_base = HEDGE_BASE.count_all(morphs, raw)["완화"]["count"]
        rows = extract(raw)
        out[c] = {"eoj": eoj, "booster": booster, "h_base": h_base,
                  "eng_suitda": h_full - h_base, "rows": rows, "regex": len(rows)}
    return out


def build(data, path):
    wb = Workbook()
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name=KO, bold=True, color="FFFFFF", size=10)
    need_fill = PatternFill("solid", fgColor="FFF2CC")     # 입력 필요(노랑)
    blue = Font(name=KO, color="0000FF")                   # 하드코딩 입력
    black = Font(name=KO, color="000000")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    # ── 시트1: 코딩 지침 ──
    ws0 = wb.active
    ws0.title = "지침"
    guide = [
        ("'수 있다' 완화 판정 — 코딩 지침", 14, True),
        ("", 10, False),
        ("목적: '~ㄹ 수 있다'가 화자의 명제 확신을 '완화(hedge)'하는지 판정한다.", 10, False),
        ("BHR(강화·완화 비율)에서 이 표현이 완화 집계를 좌우하므로, 용례별 의미 구분이 필요하다.", 10, False),
        ("", 10, False),
        ("[판정 기준]  각 용례를 다음 중 하나로 코딩한다(드롭다운).", 11, True),
        ("  · 완화   — 화자가 명제 확신을 누그러뜨림. '아마/~일지도'로 바꿔도 뜻이 통한다.", 10, False),
        ("            예) '~라고 할 수 있다'(판단 유보), '그럴 수 있다', '~일 수 있다'", 10, False),
        ("  · 비완화 — 완화가 아님. 두 갈래.", 10, False),
        ("            (능력) 행위 주체의 능력·수행: '표현할 수 있다'='할 능력이 있다'", 10, False),
        ("            (제시) 증거 제시·관찰: '~을 볼 수 있다'='~이 관찰된다'(사실 진술에 가까움)", 10, False),
        ("  · 보류   — 문맥만으로 판단 불가. (집계에서 비완화로 처리되니, 가급적 결정 권장)", 10, False),
        ("", 10, False),
        ("[제시적 '볼 수 있다'가 쟁점]", 11, True),
        ("  '우리는 X를 볼 수 있다'를 완화로 볼지(독자 관여) 비완화로 볼지(사실 제시)는 이론이 갈린다.", 10, False),
        ("  당신의 판단이 곧 본 연구의 코딩 기준이 된다. 일관되게 적용하라.", 10, False),
        ("", 10, False),
        ("[사용법]", 11, True),
        ("  1. [코딩] 시트에서 각 행의 앞·뒤 문맥을 읽고 H열 '판정'에 드롭다운으로 입력.", 10, False),
        ("  2. '자동분류(참고)'는 거친 휴리스틱이니 맹신하지 말 것 — 직접 판단이 목적이다.", 10, False),
        ("  3. [집계] 시트에서 BHR이 실시간 재계산된다. '셋 다 <1'이 유지되는지 관찰.", 10, False),
        ("  4. 모두 '완화'로 코딩하면 현행 게재값(R0)과 일치한다(검산용).", 10, False),
        ("", 10, False),
        ("[주의] 단일 코더 판정은 본문 근거로 부족하다. 제2 코더와의 일치율(κ)을 별도 보고할 것.", 10, True),
    ]
    for i, (txt, sz, bold) in enumerate(guide, 1):
        cell = ws0.cell(row=i, column=1, value=txt)
        cell.font = Font(name=KO, size=sz, bold=bold,
                         color="1F4E78" if bold and sz >= 12 else "000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 100

    # ── 시트2: 코딩 ──
    ws = wb.create_sheet("코딩")
    headers = ["ID", "비평가", "앞 문맥", "「수 있-」", "뒤 문맥",
               "앞 서술어", "자동분류(참고)", "판정", "메모"]
    widths = [5, 7, 50, 13, 50, 10, 13, 9, 18]
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for critic in CRITICS:
        for (left, phrase, right, prev, auto) in data[critic]["rows"]:
            vals = [r - 1, critic, left, phrase, right, prev, AUTOLABEL[auto], None, None]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.border = border
                cell.font = Font(name=KO, size=10)
                if j in (3, 5, 9):
                    cell.alignment = wrap
                elif j == 4:
                    cell.alignment = center
                    cell.font = Font(name=KO, size=10, bold=True)
                else:
                    cell.alignment = center
                if j == 8:
                    cell.fill = need_fill
            ws.row_dimensions[r].height = 30
            r += 1
    last = r - 1
    ws.freeze_panes = "A2"
    dv = DataValidation(type="list", formula1='"완화,비완화,보류"', allow_blank=True)
    dv.add(f"H2:H{last}")
    ws.add_data_validation(dv)

    # ── 시트3: 집계(BHR) ──
    wsa = wb.create_sheet("집계(BHR)")
    title = wsa.cell(row=1, column=1, value="BHR 실시간 재집계 — [코딩] 시트 판정에 연동")
    title.font = Font(name=KO, bold=True, size=13, color="1F4E78")
    note = wsa.cell(row=2, column=1,
                    value="파란 글씨=파이프라인 상수(고정 입력) · 검은 글씨=수식. 모두 '완화'면 BHR=현행값.")
    note.font = Font(name=KO, size=9, italic=True, color="808080")

    cols = ["비평가", "강조(raw)", "비수있다완화", "수있다(엔진)", "수있다(코딩총)",
            "어절", "코딩'완화'수", "완화합계", "강조NF", "완화NF", "BHR"]
    cw = [8, 11, 13, 12, 13, 9, 12, 11, 10, 10, 9]
    hr = 4
    for j, (h, w) in enumerate(zip(cols, cw), 1):
        c = wsa.cell(row=hr, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        wsa.column_dimensions[get_column_letter(j)].width = w

    # 코딩 시트 범위
    CR = f"'코딩'!$B$2:$B${last}"
    HR = f"'코딩'!$H$2:$H${last}"
    for i, critic in enumerate(CRITICS):
        row = hr + 1 + i
        d = data[critic]
        inputs = {1: critic, 2: d["booster"], 3: d["h_base"],
                  4: d["eng_suitda"], 5: d["regex"], 6: d["eoj"]}
        for col, val in inputs.items():
            c = wsa.cell(row=row, column=col, value=val)
            c.border = border; c.alignment = center
            c.font = Font(name=KO) if col == 1 else blue
        # G: 코딩 '완화' 수 (COUNTIFS)
        g = wsa.cell(row=row, column=7,
                     value=f'=COUNTIFS({CR},$A{row},{HR},"완화")')
        # H: 완화 합계 = 비수있다완화 + 수있다엔진 × (코딩완화수 / 코딩총)
        h = wsa.cell(row=row, column=8,
                     value=f"=$C{row}+$D{row}*($G{row}/$E{row})")
        # I: 강조NF, J: 완화NF, K: BHR (파이프라인 _bhr 와 동일한 반올림)
        bnf = wsa.cell(row=row, column=9, value=f"=ROUND($B{row}/$F{row}*1000,2)")
        hnf = wsa.cell(row=row, column=10, value=f"=ROUND($H{row}/$F{row}*1000,2)")
        bhr = wsa.cell(row=row, column=11, value=f"=ROUND($I{row}/$J{row},3)")
        for col in (7, 8, 9, 10, 11):
            cc = wsa.cell(row=row, column=col)
            cc.border = border; cc.alignment = center; cc.font = black
        bnf.number_format = "0.00"; hnf.number_format = "0.00"
        bhr.number_format = "0.000"; h.number_format = "0.0"
        bhr.font = Font(name=KO, bold=True)

    # 판정 지표
    k0, k1, k2 = hr + 1, hr + 2, hr + 3   # 박/임/김 BHR 행
    ind = hr + 5
    wsa.cell(row=ind, column=1, value="결론 점검").font = Font(name=KO, bold=True, size=11)
    wsa.cell(row=ind + 1, column=1, value="셋 다 BHR<1 (register 수렴)").font = Font(name=KO)
    wsa.cell(row=ind + 1, column=4,
             value=f'=IF(MAX(K{k0}:K{k2})<1,"성립 (셋 다 <1)","깨짐 (≥1 존재)")').font = Font(name=KO, bold=True)
    wsa.cell(row=ind + 2, column=1, value="박용철 최저 BHR (영감-유보)").font = Font(name=KO)
    wsa.cell(row=ind + 2, column=4,
             value=f'=IF(AND(K{k0}<K{k1},K{k0}<K{k2}),"성립 (박 최저)","서열 변동")').font = Font(name=KO, bold=True)
    wsa.cell(row=ind + 4, column=1,
             value="참고 — 현행 R0(전부 완화) BHR: 박용철 0.296 · 임화 0.609 · 김기림 0.583").font = \
        Font(name=KO, size=9, color="808080")
    wsa.column_dimensions["D"].width = 22

    wb.save(path)
    return last


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--corpus", default=str(Path(__file__).resolve().parents[2]))
    ap.add_argument("--sw-tokenizer", default=None)
    ap.add_argument("--no-download", action="store_true")
    ap.add_argument("--output", default="output")
    a = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    corpus = build_corpus(a.corpus, sw_model_path=a.sw_tokenizer,
                          auto_download=not a.no_download)
    data = per_critic_constants(corpus)
    out = Path(a.output); out.mkdir(parents=True, exist_ok=True)
    path = out / "suitda_coding.xlsx"
    n = build(data, path)
    print(f"✓ 저장: {path}")
    for c in CRITICS:
        d = data[c]
        print(f"  {c}: 용례 {d['regex']}개 · 강조 {d['booster']} · "
              f"비수있다완화 {d['h_base']} · 수있다(엔진) {d['eng_suitda']} · 어절 {d['eoj']}")
    print(f"  코딩 행 {n-1}개. [코딩] 시트 H열에 완화/비완화/보류 입력 → [집계] 시트 BHR 자동.")


if __name__ == "__main__":
    main()
