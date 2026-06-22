#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_bhr_mullon.py — '물론'을 구문 규칙으로 강조/비강조 구분 후 BHR 재집계 + 코딩 워크시트

'ㄹ 수 있다'(완충)와 대칭으로, 강조 1위 '물론'을 의미 구분한다.
KWIC 전수 검토 결과 이 비평 레지스터의 '물론'은 대부분 양보·첨가·'물론하고'(=regardless)로,
화자가 자기 명제를 단언하는 강조가 아니라 반대 논점을 인정한 뒤 뒤집는 구조다.

규칙(휴리스틱, 수작업 검증 대상):
  강조(booster) : '물론이다/물론이었다/물론이지/물론이오' 등 확신 술어 ('것은 물론이다')
  비강조        : 그 외 — 양보 부사('물론 X 그러나 Y'), '물론하고'(regardless),
                  첨가('A는 물론 B도', '물론이어니와/물론이요'), 양보 종결('물론이나')

산출:
  · R0(전부완화·물론강조) / 형태완충(수있다 의미구분) / 양측정제(수있다+물론) BHR 비교
  · output/mullon_coding.xlsx — 전체 '물론' 용례 + 자동분류 + 제2 코더 판정칸

  python run_bhr_mullon.py
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from gigyo import build_corpus, config as C
from gigyo.core import BOOSTER_ENGINE, MarkerEngine

CRITICS = ["박용철", "임화", "김기림"]
SUITDA = re.compile(r'([가-힣]+)\s*수(가|도|는|를|만)?\s*(있[가-힣]*)')
HEDGE_END = re.compile(r'(으나|지마|지만|으리|을지|는지|을까|을런지|음직)')
MULLON = re.compile(r'(?<!유)물론')                      # '유물론' 등 복합어 제외
BOOST_PRED = re.compile(r'\s?이(다|었|겠|지|오|에)')     # '물론이다'류 확신 술어

HEDGE_FULL = MarkerEngine(C.HEDGES_INV)
HEDGE_NOSU = MarkerEngine({"완화": [e for e in C.HEDGES_INV["완화"]
                                    if e["label"] != "ㄹ 수 있다"]})


def suitda_hedge_count(raw):
    flat = re.sub(r"\s+", " ", raw)
    return sum(1 for m in SUITDA.finditer(flat)
               if (m.group(2) in ("는", "도")) or HEDGE_END.search(m.group(3)))


def classify_mullon(raw):
    flat = re.sub(r"\s+", " ", raw)
    boost = nonb = 0
    rows = []
    for m in MULLON.finditer(flat):
        after = flat[m.end():m.end() + 6]
        is_b = bool(BOOST_PRED.match(after))
        boost += is_b
        nonb += (not is_b)
        s, e = max(0, m.start() - 32), min(len(flat), m.end() + 50)
        rows.append((flat[s:m.start()] + "【물론】" + flat[m.end():e],
                     "강조" if is_b else "비강조"))
    return boost, nonb, rows


def bhr(b, h, eoj):
    bn = round(b / eoj * 1000, 2)
    hn = round(h / eoj * 1000, 2)
    return round(bn / hn, 3) if hn else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--corpus", default=str(Path(__file__).resolve().parents[2]))
    ap.add_argument("--sw-tokenizer", default=None)
    ap.add_argument("--no-download", action="store_true")
    ap.add_argument("--output", default="output")
    a = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    corpus = build_corpus(a.corpus, sw_model_path=a.sw_tokenizer,
                          auto_download=not a.no_download)

    data, dump = {}, []
    for c in CRITICS:
        raw = corpus.critic_raw_text(c)
        eoj = corpus.critic_eojeols(c)
        morphs = corpus.tokenizer.morphs(raw)
        b_full = BOOSTER_ENGINE.count_all(morphs, raw)["강조"]["count"]
        h_full = HEDGE_FULL.count_all(morphs, raw)["완화"]["count"]            # R0 완충
        h_form = HEDGE_NOSU.count_all(morphs, raw)["완화"]["count"] + suitda_hedge_count(raw)
        mb, mnb, rows = classify_mullon(raw)
        data[c] = {"eoj": eoj, "b_full": b_full, "h_full": h_full, "h_form": h_form,
                   "m_boost": mb, "m_nonb": mnb, "b_adj": b_full - mnb}
        for kwic, auto in rows:
            dump.append((c, kwic, auto))

    sep = "=" * 78
    print(sep, "\n  '물론' 구문 분류 (전수; 휴리스틱·수작업 검증 대상)\n", sep, sep="")
    print(f"  {'비평가':<6}{'물론총':>7}{'강조(유지)':>10}{'비강조(제외)':>12}")
    for c in CRITICS:
        d = data[c]
        print(f"  {c:<6}{d['m_boost']+d['m_nonb']:>7}{d['m_boost']:>10}{d['m_nonb']:>12}")

    print("\n", sep, "\n  BHR 3단 비교\n", sep, sep="")
    print(f"  {'비평가':<6}{'R0(전부완화)':>13}{'형태완충(수있다)':>16}{'양측정제(+물론)':>16}")
    res = {}
    for c in CRITICS:
        d = data[c]
        r0 = bhr(d["b_full"], d["h_full"], d["eoj"])
        form = bhr(d["b_full"], d["h_form"], d["eoj"])
        both = bhr(d["b_adj"], d["h_form"], d["eoj"])
        res[c] = both
        print(f"  {c:<6}{r0:>13.3f}{form:>16.3f}{both:>16.3f}")

    below = all(v < 1 for v in res.values())
    park = all(res["박용철"] < res[o] for o in CRITICS if o != "박용철")
    print(f"\n  [양측 정제 결론]  셋 다 <1: {'성립' if below else '깨짐'}"
          f"   |  서열 임>김>박: {'성립' if res['임화']>res['김기림']>res['박용철'] else '변동'}"
          f"   |  박용철 최저: {'성립' if park else '아님'}")
    print("  → 완충(수있다)과 강조(물론)를 대칭으로 정제하면 임화도 1 미만(단언 주장 폐기),")
    print("     코딩 불변 결론은 상대 서열(임>김>박·박용철 최저)뿐.")

    # ── 코딩 워크시트 ──
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("\n  (openpyxl 없음 — 워크시트 생략)")
        return
    KO = "맑은 고딕"
    wb = Workbook(); ws = wb.active; ws.title = "물론"
    thin = Side(style="thin", color="BBBBBB"); bd = Border(thin, thin, thin, thin)
    hdr = ["ID", "비평가", "문맥 (【물론】)", "자동분류", "판정", "메모"]
    wdt = [4, 7, 95, 9, 9, 18]
    for j, (h, w) in enumerate(zip(hdr, wdt), 1):
        cl = ws.cell(1, j, h); cl.font = Font(name=KO, bold=True, color="FFFFFF")
        cl.fill = PatternFill("solid", fgColor="1F4E78"); cl.alignment = Alignment("center", "center")
        ws.column_dimensions[chr(64+j)].width = w
    for i, (c, kwic, auto) in enumerate(dump, 2):
        vals = [i-1, c, kwic, auto, None, None]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(i, j, v); cell.border = bd; cell.font = Font(name=KO, size=10)
            cell.alignment = Alignment(wrap_text=(j == 3), vertical="top",
                                       horizontal="left" if j == 3 else "center")
            if j == 5:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.row_dimensions[i].height = 30
    ws.freeze_panes = "A2"
    dv = DataValidation(type="list", formula1='"강조,비강조,보류"', allow_blank=True)
    dv.add(f"E2:E{len(dump)+1}"); ws.add_data_validation(dv)
    out = Path(a.output); out.mkdir(parents=True, exist_ok=True)
    p = out / "mullon_coding.xlsx"; wb.save(p)
    print(f"\n  코딩 워크시트 저장(제2 코더용): {p}  (총 {len(dump)}용례)")


if __name__ == "__main__":
    main()
