#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_sample_coding.py — '수 있다' 40개 표본 코딩지(문단 문맥 포함)

전수(255) 대신 비평가별 층화 무작위 표본 40개만 코딩한다(자기 진단용).
각 용례에 '해당 문단 전체'를 붙여 의미 파악이 쉽도록 한다.
판정을 입력하면 [집계] 시트가 표본 완화비율 → BHR 추정을 실시간 계산한다.

  · 표본: 임화 20 · 김기림 10 · 박용철 10 (시드 고정 → 재현)
  · 경계 정확성 유지: 표본을 전부 '완화'로 코딩하면 R0(현행값), 전부 '비완화'면 R3.
    중간값만 표본 추정(소표본이라 불확실; 본문 검증이 아니라 밴드 위치 진단용).

  python make_sample_coding.py            # → output/suitda_sample40.xlsx
"""
from __future__ import annotations

import argparse
import random
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from gigyo import build_corpus
from make_coding_sheet import per_critic_constants, KO, AUTOLABEL
from run_bhr_sense import PAT, classify, CRITICS

SAMPLE_N = {"임화": 20, "김기림": 10, "박용철": 10}
SEED = 42


def extract_paragraphs(corpus):
    """corpus.raw_texts 를 줄(문단) 단위로 쪼개 '수 있-' 매치를 문단째 추출."""
    recs = {c: [] for c in CRITICS}
    for (critic, stem), raw in corpus.raw_texts.items():
        if critic not in recs:
            continue
        for line in raw.split("\n"):
            para = line.strip()
            if not para:
                continue
            for m in PAT.finditer(para):
                prev = m.group(1)
                pre = para[max(0, m.start() - 20):m.start()].strip().split(" ")
                prev2 = pre[-1] if pre else ""
                marked = para[:m.start()] + "【" + m.group(0) + "】" + para[m.end():]
                recs[critic].append({
                    "stem": stem, "para": marked, "phrase": m.group(0),
                    "prev": prev, "auto": classify(prev, prev2)})
    return recs


def build(sample, consts, path):
    wb = Workbook()
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name=KO, bold=True, color="FFFFFF", size=10)
    need_fill = PatternFill("solid", fgColor="FFF2CC")
    blue = Font(name=KO, color="0000FF")
    black = Font(name=KO, color="000000")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    # ── 지침 ──
    ws0 = wb.active; ws0.title = "지침"
    guide = [
        ("'수 있다' 완화 판정 — 40개 표본 코딩지 (자기 진단용)", 14, True),
        ("", 10, False),
        ("이것은 전수 검증이 아니라, '의미를 제대로 구분하면 BHR이 밴드 어디쯤 떨어지는가'를", 10, False),
        ("가늠하기 위한 층화 무작위 표본(임화 20·김기림 10·박용철 10, 시드 42)이다.", 10, False),
        ("본문 근거로 쓰려면 제2 코더 + 일치율(κ)이 따로 필요하다(이 표본만으론 부족).", 10, True),
        ("", 10, False),
        ("[판정 기준]  각 용례를 완화 / 비완화 / 보류 중 하나로(드롭다운).", 11, True),
        ("  · 완화   — 화자가 명제 확신을 누그러뜨림. '아마/~일지도'로 바꿔도 통한다.", 10, False),
        ("            예) '~라고 할 수 있다'(판단 유보), '그럴 수 있다', '~일 수 있다'", 10, False),
        ("  · 비완화 — (능력) '표현할 수 있다'=할 능력이 있다 / (제시) '~을 볼 수 있다'=관찰된다(사실 진술)", 10, False),
        ("  · 보류   — 문맥으로도 판단 불가(집계에서 비완화로 처리되니 가급적 결정).", 10, False),
        ("", 10, False),
        ("[쟁점] '볼 수 있다'(제시)를 완화로 볼지 비완화로 볼지는 이론이 갈린다.", 11, True),
        ("  당신의 일관된 입장이 곧 코딩 기준이 된다. 문단 전체를 읽고 판단하라.", 10, False),
        ("", 10, False),
        ("[사용법]", 11, True),
        ("  1. [표본40] 시트의 '문단' 열을 읽고(【 】안이 대상 표현) H열 '판정'에 입력.", 10, False),
        ("  2. [집계] 시트에서 표본 완화비율로 BHR 추정이 실시간 갱신된다.", 10, False),
        ("  3. 전부 완화→R0(현행), 전부 비완화→R3. 경계는 정확, 중간은 표본 추정(소표본 주의).", 10, False),
    ]
    for i, (txt, sz, bold) in enumerate(guide, 1):
        c = ws0.cell(row=i, column=1, value=txt)
        c.font = Font(name=KO, size=sz, bold=bold,
                      color="1F4E78" if bold and sz >= 12 else ("C00000" if bold else "000000"))
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 105

    # ── 표본40 ──
    ws = wb.create_sheet("표본40")
    headers = ["ID", "비평가", "출처(텍스트)", "문단 (【 】=판정 대상)",
               "「수 있-」", "앞 서술어", "자동분류(참고)", "판정", "메모"]
    widths = [4, 7, 24, 92, 13, 9, 13, 9, 16]
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for critic in CRITICS:                       # 박/임/김 순
        for rec in sample[critic]:
            vals = [r - 1, critic, rec["stem"], rec["para"], rec["phrase"],
                    rec["prev"], AUTOLABEL[rec["auto"]], None, None]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.border = border; cell.font = Font(name=KO, size=10)
                if j == 4:
                    cell.alignment = wrap
                elif j == 9:
                    cell.alignment = wrap
                elif j == 5:
                    cell.alignment = center; cell.font = Font(name=KO, size=10, bold=True)
                else:
                    cell.alignment = center
                if j == 8:
                    cell.fill = need_fill
            ws.row_dimensions[r].height = 95
            r += 1
    last = r - 1
    ws.freeze_panes = "A2"
    dv = DataValidation(type="list", formula1='"완화,비완화,보류"', allow_blank=True)
    dv.add(f"H2:H{last}")
    ws.add_data_validation(dv)

    # ── 집계(BHR 표본추정) ──
    wsa = wb.create_sheet("집계(BHR추정)")
    wsa.cell(row=1, column=1, value="BHR 표본추정 — [표본40] 판정에 연동").font = \
        Font(name=KO, bold=True, size=13, color="1F4E78")
    wsa.cell(row=2, column=1,
             value="표본 완화비율을 '수 있다' 전체에 적용한 추정. 경계(전부완화/전부비완화)는 정확, 중간은 소표본 추정.").font = \
        Font(name=KO, size=9, italic=True, color="808080")
    cols = ["비평가", "강조", "비수있다완화", "수있다(엔진)", "어절",
            "표본수", "표본'완화'", "완화비율", "완화합계", "강조NF", "완화NF", "BHR추정"]
    cw = [8, 7, 12, 12, 9, 8, 9, 9, 11, 9, 9, 10]
    hr = 4
    for j, (h, w) in enumerate(zip(cols, cw), 1):
        c = wsa.cell(row=hr, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        wsa.column_dimensions[get_column_letter(j)].width = w
    CR = f"'표본40'!$B$2:$B${last}"
    HR = f"'표본40'!$H$2:$H${last}"
    for i, critic in enumerate(CRITICS):
        row = hr + 1 + i
        d = consts[critic]
        for col, val in {1: critic, 2: d["booster"], 3: d["h_base"],
                         4: d["eng_suitda"], 5: d["eoj"], 6: len(sample[critic])}.items():
            c = wsa.cell(row=row, column=col, value=val)
            c.border = border; c.alignment = center
            c.font = Font(name=KO) if col == 1 else blue
        wsa.cell(row=row, column=7, value=f'=COUNTIFS({CR},$A{row},{HR},"완화")')
        wsa.cell(row=row, column=8, value=f"=IF($F{row}=0,0,$G{row}/$F{row})").number_format = "0.0%"
        wsa.cell(row=row, column=9, value=f"=$C{row}+$D{row}*$H{row}").number_format = "0.0"
        wsa.cell(row=row, column=10, value=f"=ROUND($B{row}/$E{row}*1000,2)").number_format = "0.00"
        wsa.cell(row=row, column=11, value=f"=ROUND($I{row}/$E{row}*1000,2)").number_format = "0.00"
        bhr = wsa.cell(row=row, column=12, value=f"=ROUND($J{row}/$K{row},3)")
        bhr.number_format = "0.000"; bhr.font = Font(name=KO, bold=True)
        for col in range(7, 12):
            wsa.cell(row=row, column=col).border = border
            wsa.cell(row=row, column=col).alignment = center
            if wsa.cell(row=row, column=col).font.color is None:
                wsa.cell(row=row, column=col).font = black
        bhr.border = border; bhr.alignment = center

    k0, k2 = hr + 1, hr + 3
    ind = hr + 5
    wsa.cell(row=ind, column=1, value="결론 점검(추정)").font = Font(name=KO, bold=True, size=11)
    wsa.cell(row=ind + 1, column=1, value="셋 다 BHR<1 (register 수렴)").font = Font(name=KO)
    wsa.cell(row=ind + 1, column=4,
             value=f'=IF(MAX(L{k0}:L{k2})<1,"성립(셋 다 <1)","깨짐(≥1 존재)")').font = Font(name=KO, bold=True)
    wsa.cell(row=ind + 2, column=1, value="박용철 최저 BHR").font = Font(name=KO)
    wsa.cell(row=ind + 2, column=4,
             value=f'=IF(AND(L{k0}<L{k0+1},L{k0}<L{k2}),"성립(박 최저)","서열 변동")').font = Font(name=KO, bold=True)
    wsa.cell(row=ind + 4, column=1,
             value="참고 — R0(전부완화): 박0.296·임0.609·김0.583 | R3(전부비완화): 박0.515·임1.255·김0.891").font = \
        Font(name=KO, size=9, color="808080")
    wsa.column_dimensions["D"].width = 20

    wb.save(path)
    return last


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--corpus", default=str(Path(__file__).resolve().parents[2]))
    ap.add_argument("--sw-tokenizer", default=None)
    ap.add_argument("--no-download", action="store_true")
    ap.add_argument("--output", default="output")
    a = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    corpus = build_corpus(a.corpus, sw_model_path=a.sw_tokenizer,
                          auto_download=not a.no_download)
    consts = per_critic_constants(corpus)
    recs = extract_paragraphs(corpus)

    rng = random.Random(SEED)
    sample = {}
    for c in CRITICS:
        pool = recs[c]
        n = min(SAMPLE_N[c], len(pool))
        sample[c] = rng.sample(pool, n)

    out = Path(a.output); out.mkdir(parents=True, exist_ok=True)
    path = out / "suitda_sample40.xlsx"
    last = build(sample, consts, path)
    print(f"✓ 저장: {path}  (코딩 행 {last-1}개)")
    for c in CRITICS:
        ids = ", ".join(sorted({s["stem"] for s in sample[c]}))
        print(f"  {c}: 표본 {len(sample[c])}개 / 전체 {len(recs[c])}개  ← {ids}")
    print("  [표본40] H열 코딩 → [집계] BHR추정 자동. 전부완화=R0, 전부비완화=R3로 검산.")


if __name__ == "__main__":
    main()
