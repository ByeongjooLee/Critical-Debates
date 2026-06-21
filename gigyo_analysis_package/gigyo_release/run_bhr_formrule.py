#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_bhr_formrule.py — '수 있다' 형태 기반 하위유형 규칙으로 BHR 전수 재집계 + 수작업 교차검증

규칙(재현 가능·자동):
  완화(hedge)  = 양보·추측 형태
                 · 보조사 '는/도'  ('-ㄹ 수는/수도 있-')
                 · 추측·양보 어미  (있으나/있지만/있지마/있으리/있을지/있는지/있을까/있을런지/있음직)
  비완화        = 그 외 평서·관형형 (제시적·능력 용법; 예 '볼 수 있다', '파악할 수 있다')

근거: 본 비평 레지스터에서 '-ㄹ 수 있다'는 평서·관형형(지각·발견 술어)으로 나타나는
제시적 용법이 지배적이며, 진정한 완화는 양보·추측 형태로 형식상 식별되는 소수에 국한.
일반 한국어 완화 목록(전부 완화)은 비평 산문에서 완화를 과대 집계함.

  python run_bhr_formrule.py                 # 전수 재집계 + 40개 교차검증
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from gigyo import build_corpus, config as C
from gigyo.core import BOOSTER_ENGINE, MarkerEngine

CRITICS = ["박용철", "임화", "김기림"]
PAT = re.compile(r'([가-힣]+)\s*수(가|도|는|를|만)?\s*(있[가-힣]*)')
HEDGE_END = re.compile(r'(으나|지마|지만|으리|을지|는지|을까|을런지|음직)')

HEDGE_BASE = MarkerEngine({cat: [e for e in es if e["label"] != "ㄹ 수 있다"]
                           for cat, es in C.HEDGES_INV.items()})
HEDGE_FULL = MarkerEngine(C.HEDGES_INV)


def form_is_hedge(particle: str, euot: str) -> bool:
    """양보·추측 형태면 완화."""
    if particle in ("는", "도"):
        return True
    return bool(HEDGE_END.search(euot))


def classify_corpus(corpus):
    """비평가별 '수 있다' 형태 분류 + 엔진 상수."""
    out = {}
    for c in CRITICS:
        raw = corpus.critic_raw_text(c)
        flat = re.sub(r"\s+", " ", raw)
        hedge = nonh = 0
        for m in PAT.finditer(flat):
            if form_is_hedge(m.group(2) or "", m.group(3)):
                hedge += 1
            else:
                nonh += 1
        morphs = corpus.tokenizer.morphs(raw)
        booster = BOOSTER_ENGINE.count_all(morphs, raw)["강조"]["count"]
        h_full = HEDGE_FULL.count_all(morphs, raw)["완화"]["count"]
        h_base = HEDGE_BASE.count_all(morphs, raw)["완화"]["count"]
        out[c] = {"eoj": corpus.critic_eojeols(c), "booster": booster,
                  "h_base": h_base, "eng_suitda": h_full - h_base,
                  "regex": hedge + nonh, "hedge_form": hedge, "nonhedge_form": nonh}
    return out


def bhr_from(booster, hedge, eoj):
    b_nf = round(booster / eoj * 1000, 2)
    h_nf = round(hedge / eoj * 1000, 2)
    return round(b_nf / h_nf, 3) if h_nf > 0 else None


def bhr_variants(d):
    """R0(전부완화)·형태규칙·R3(전부비완화) BHR."""
    f = d["hedge_form"] / d["regex"] if d["regex"] else 0
    return {
        "R0": bhr_from(d["booster"], d["h_base"] + d["eng_suitda"], d["eoj"]),
        "form": bhr_from(d["booster"], d["h_base"] + d["eng_suitda"] * f, d["eoj"]),
        "R3": bhr_from(d["booster"], d["h_base"], d["eoj"]),
    }


def crossval(path):
    """수작업 40개 vs 형태규칙 일치율."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    if not Path(path).exists():
        return None
    wb = load_workbook(path, data_only=True)
    ws = wb["표본40"]
    agree = disagree = boryu = 0
    rows = []
    for r in range(2, ws.max_row + 1):
        critic = ws.cell(row=r, column=2).value
        phrase = ws.cell(row=r, column=5).value or ""
        human = ws.cell(row=r, column=8).value
        if not critic:
            continue
        m = PAT.search(phrase.replace(" ", ""))
        rule = "완화" if (m and form_is_hedge(m.group(2) or "", m.group(3))) else "비완화"
        if human == "보류":
            boryu += 1
            continue
        if human == rule:
            agree += 1
        else:
            disagree += 1
            rows.append((critic, phrase, human, rule))
    return {"agree": agree, "disagree": disagree, "boryu": boryu, "diffs": rows}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--corpus", default=str(Path(__file__).resolve().parents[2]))
    ap.add_argument("--sw-tokenizer", default=None)
    ap.add_argument("--no-download", action="store_true")
    ap.add_argument("--coded", default="output/suitda_sample40_me.xlsx")
    a = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    corpus = build_corpus(a.corpus, sw_model_path=a.sw_tokenizer,
                          auto_download=not a.no_download)
    data = classify_corpus(corpus)

    sep = "=" * 76
    print(sep, "\n  '수 있다' 형태 기반 하위유형 분류 (전수)\n", sep, sep="")
    print(f"  {'비평가':<6}{'총':>5}{'완화형':>7}{'비완화형':>8}{'완화비율':>9}")
    for c in CRITICS:
        d = data[c]
        f = d["hedge_form"] / d["regex"] if d["regex"] else 0
        print(f"  {c:<6}{d['regex']:>5}{d['hedge_form']:>7}{d['nonhedge_form']:>8}{f*100:>8.0f}%")

    print("\n", sep, "\n  BHR — R0(전부완화) vs 형태규칙 vs R3(전부비완화)\n", sep, sep="")
    print(f"  {'비평가':<6}{'R0':>8}{'형태규칙':>9}{'R3':>8}")
    form = {}
    for c in CRITICS:
        v = bhr_variants(data[c]); form[c] = v["form"]
        print(f"  {c:<6}{v['R0']:>8.3f}{v['form']:>9.3f}{v['R3']:>8.3f}")

    below = all(form[c] < 1 for c in CRITICS)
    park_low = all(form["박용철"] < form[o] for o in CRITICS if o != "박용철")
    print(f"\n  [형태규칙 결론]  셋 다 BHR<1: {'성립' if below else '깨짐(임화 등 ≥1)'}"
          f"   |  박용철 최저: {'성립' if park_low else '아님'}")

    cv = crossval(a.coded)
    print("\n", sep, "\n  수작업 40개 vs 형태규칙 — 교차검증\n", sep, sep="")
    if cv is None:
        print("  (코딩 파일 없음 — 건너뜀)")
    else:
        tot = cv["agree"] + cv["disagree"]
        print(f"  일치 {cv['agree']}/{tot} ({cv['agree']/tot*100:.0f}%)  "
              f"불일치 {cv['disagree']}  보류 {cv['boryu']}")
        if cv["diffs"]:
            print("  [불일치]")
            for critic, ph, human, rule in cv["diffs"]:
                print(f"     {critic} '{ph}'  수작업={human} / 규칙={rule}")
        else:
            print("  → 형태규칙이 인간 판정을 100% 복제. 규칙의 재현성·타당성 입증.")


if __name__ == "__main__":
    main()
